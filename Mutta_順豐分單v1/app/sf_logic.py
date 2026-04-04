"""
分單系統（順豐模式）核心邏輯
- 宅配 PDF：品項在 y≈220~450（無 7-11 標籤）
- 分類邏輯與分單系統完全相同（共用 core_logic.py）
"""
import re, os, sys, copy
from collections import defaultdict

sys.path.insert(0, os.path.dirname(__file__))
from core_logic import (
    N, gender, parse_from_items, gift, special_summary,
    inject_price_map, inject_gift_rules,
    render_label, add_image_to_page,
)
from pdfplumber import open as pdf_open
from pypdf import PdfReader, PdfWriter


def extract_sf_line_items(page, y_min=220, y_max=500):
    """宅配 PDF 品項提取（動態偵測欄位位置，支援有無商品圖片欄）"""
    ws = page.extract_words()

    seq_rows = sorted(
        [(w['top'], int(w['text'])) for w in ws
         if 35 <= w['x0'] <= 55 and y_min < w['top'] < y_max
         and re.match(r'^\d+$', w['text'])],
        key=lambda x: x[0]
    )
    if not seq_rows:
        return []

    # 動態偵測欄位 x 座標
    def find_header_x(keywords, default):
        candidates = [w for w in ws
                      if any(kw in w['text'] for kw in keywords)
                      and w['top'] < y_min + 50]
        if candidates:
            return max(candidates, key=lambda w: w['top'])['x0']
        return default

    spec_col_x  = find_header_x(['規格'],     250)
    qty_col_x   = find_header_x(['訂購數量'],  390)
    price_col_x = find_header_x(['單筆金額'],  430)

    items = []
    for idx, (seq_top, seq_num) in enumerate(seq_rows):
        t_min = seq_top - 15
        t_max = seq_rows[idx+1][0] - 5 if idx+1 < len(seq_rows) else seq_top + 80

        price_words = [w for w in ws
                       if price_col_x - 10 <= w['x0'] <= price_col_x + 60 and t_min < w['top'] < t_max
                       and re.match(r'^\$[\d,]+$', w['text'])]
        if not price_words:
            continue
        try:
            price = int(price_words[0]['text'].replace('$','').replace(',',''))
        except ValueError:
            continue

        qty_words = [w for w in ws
                     if qty_col_x - 5 <= w['x0'] <= qty_col_x + 30 and t_min < w['top'] < t_max
                     and re.match(r'^\d+$', w['text'])]
        qty = int(qty_words[0]['text']) if qty_words else 1

        spec_words = sorted(
            [w for w in ws if spec_col_x - 5 <= w['x0'] <= qty_col_x - 5 and t_min < w['top'] < t_max],
            key=lambda w: (w['top'], w['x0'])
        )
        spec_raw = N(' '.join(w['text'] for w in spec_words))

        items.append({'price': price, 'qty': qty, 'spec_raw': spec_raw})

    return items


def parse_sf_pdf(pdf_path: str, csv_data: dict) -> list:
    """
    解析宅配 PDF，回傳訂單清單
    csv_data: {訂單編號: {'name': ..., 'phone': ..., 'address': ...}}
    """
    orders = []
    with pdf_open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        i = 0
        while i < total_pages:
            page = pdf.pages[i]
            text = page.extract_text() or ''

            oid_m = re.search(r'#(\w{8,})', text)
            if not oid_m:
                i += 1
                continue
            oid = oid_m.group(1)

            buyer_m = re.search(r'收件[人⼈]：(\S+)', text)
            buyer = buyer_m.group(1) if buyer_m else ''

            disc_m = re.search(r'訂單折扣\s+-?\$?([\d,]+)', text)
            disc_val = int(disc_m.group(1).replace(',', '')) if disc_m else 0
            disc_str = f'-${disc_val}' if disc_val else ''

            # 跨頁偵測
            page_indices = [i]
            j = i + 1
            while j < total_pages:
                nxt = pdf.pages[j].extract_text() or ''
                if re.search(r'#\w{8,}', nxt):
                    break
                if re.search(r'總計|訂單金額|\$[\d,]+', nxt):
                    page_indices.append(j)
                    j += 1
                else:
                    break

            # 提取品項
            all_items = []
            for pi_idx, pi in enumerate(page_indices):
                y_min = 220 if pi_idx == 0 else 50
                all_items.extend(extract_sf_line_items(pdf.pages[pi], y_min=y_min))

            if not all_items:
                i = j
                continue

            sp  = parse_from_items(all_items)
            gnd = gender(buyer)
            cat, gft = gift(sp, gnd, disc_val)

            if disc_val > 158:
                cat = '特殊單'
                gft = f'請人工確認\n折扣 -${disc_val}'
            elif disc_val != 0 and disc_val != 50 and cat != '特殊單':
                gft = gft + f'  折{disc_str}'

            if cat == '特殊單' and all_items and disc_val <= 158:
                summary = special_summary(all_items)
                gft = f'請人工確認\n{summary}'

            csv_row = csv_data.get(oid, {})

            orders.append(dict(
                idx=i, oid=oid, buyer=buyer, gnd=gnd,
                disc_val=disc_val, disc_str=disc_str or '無',
                sp=sp, cat=cat, gift=gft, special=(cat == '特殊單'),
                page_indices=page_indices,
                sf_name=csv_row.get('name', buyer),
                sf_phone=csv_row.get('phone', ''),
                sf_address=csv_row.get('address', ''),
            ))
            i = j

    return orders


def generate_sf_pdfs(orders, pdf_in, out_dir, print_gift=True):
    """產生分類 PDF（加標籤，加數字前綴）"""
    reader = PdfReader(pdf_in)
    groups = defaultdict(list)
    for o in orders:
        groups[o['cat']].append(o)

    sorted_cats = sorted(groups.keys(), key=lambda c: (c == '特殊單', c))
    pad = len(str(len(sorted_cats)))

    output_files = []
    for num, cat in enumerate(sorted_cats, 1):
        ords = groups[cat]
        writer = PdfWriter()
        page_map = []
        for o in ords:
            first_idx = len(writer.pages)
            for pi in o['page_indices']:
                writer.add_page(copy.copy(reader.pages[pi]))
            page_map.append((o, first_idx))

        for o, writer_idx in page_map:
            ph = float(writer.pages[writer_idx].mediabox.height)
            png = render_label(o['cat'], o['gift'], o['special'],
                               print_gift=print_gift, pt_w=148, pt_h=172, dpi=150)
            add_image_to_page(writer, writer_idx, png, 432, ph - 210, 148, 172)

        safe = re.sub(r'[/\\ ]+', '_', cat).strip('_')
        prefix = str(num).zfill(pad)
        out_path = os.path.join(out_dir, f'{prefix}_{safe}.pdf')
        with open(out_path, 'wb') as f:
            writer.write(f)
        output_files.append(out_path)

    return output_files


def generate_sf_excel(orders, template_path, out_path):
    """產生順豐 Excel，依分類順序排列（特殊單排最後）"""
    import openpyxl, shutil
    shutil.copy(template_path, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb['information']

    sorted_orders = sorted(orders, key=lambda o: (o['cat'] == '特殊單', o['cat']))

    row = 3
    for o in sorted_orders:
        ws.cell(row, 1,  o['oid'])
        ws.cell(row, 17, o['sf_name'])
        ws.cell(row, 18, o['sf_phone'])
        ws.cell(row, 20, o['sf_address'])
        row += 1

    wb.save(out_path)
